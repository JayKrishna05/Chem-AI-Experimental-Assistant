import io
from typing import Any

import openpyxl

from backend.experiment.models import CanonicalExperiment


def parse_xlsx(data: bytes, source: str = "xlsx") -> list[CanonicalExperiment]:
    """Parse XLSX bytes into a list of CanonicalExperiments using basic heuristics."""
    experiments = []
    
    # Load workbook from bytes
    wb = openpyxl.load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        # Read header row
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration:
            continue
            
        if not headers:
            continue
            
        # Clean headers
        headers = [str(h).lower().strip() if h else None for h in headers]
        
        for row in rows:
            # Create a dict from headers and row
            row_dict = {}
            for i, val in enumerate(row):
                if i < len(headers) and headers[i] is not None:
                    if val is not None:
                        row_dict[headers[i]] = str(val).strip()
            
            if not row_dict:
                continue
                
            # Heuristic extraction
            reaction_type = row_dict.get("reaction_type") or row_dict.get("type")
            temperature_c = row_dict.get("temperature_c") or row_dict.get("temp") or row_dict.get("temperature")
            yield_percent = row_dict.get("yield_percent") or row_dict.get("yield") or row_dict.get("yield %")
            
            # Parse list fields (comma separated)
            def _parse_list(key: str) -> list[str]:
                val = row_dict.get(key, "")
                return [x.strip() for x in val.split(",")] if val else []

            reactants = _parse_list("reactants") or _parse_list("reactant")
            reagents = _parse_list("reagents") or _parse_list("reagent")
            catalysts = _parse_list("catalysts") or _parse_list("catalyst")
            products = _parse_list("products") or _parse_list("product")
            
            try:
                temp = float(temperature_c) if temperature_c else None
            except ValueError:
                temp = None
                
            try:
                yld = float(yield_percent) if yield_percent else None
            except (ValueError, TypeError):
                yld = None
                
            exp = CanonicalExperiment(
                source=source,
                reaction_type=reaction_type,
                reactants=reactants,
                reagents=reagents,
                catalysts=catalysts,
                products=products,
                temperature_c=temp,
                yield_percent=yld,
                raw_data=row_dict,
            )
            experiments.append(exp)
            
    wb.close()
    return experiments
